"""
MCC Admissions — Institute Profile & Bond PDF mirror.

For the configured listing URL (mcc.admissions.nic.in/.../ViewInstituteProfileDynamic.aspx),
visits every institute row, clicks View Profile and View Document (both are ASP.NET
__doPostBack links — no real URLs), captures the resulting PDF, and uploads to Drive:

    <MCC Archive>/Institute Profiles/Profile/<code> <name>.pdf
    <MCC Archive>/Institute Profiles/Bond/<code> <name>.pdf

Finally writes an Excel index at:
    <MCC Archive>/Institute Profiles/Institute_Profile_Index.xlsx
with columns:
    Institute Code | Institute Name | Profile PDF | Bond PDF
where the PDF columns are clickable links to the uploaded Drive files.

Reuses the OAuth credentials + DriveUploader from mcc_pipeline_starter.py.

Usage:
    python mcc_admissions_pipeline.py --limit 2     # smoke test on 2 rows
    python mcc_admissions_pipeline.py               # full run (~1-2 hours)
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path

from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout

from mcc_pipeline_starter import (
    DriveUploader, setup_logging, sanitize_name, load_config, PdfRecord,
)

LISTING_URL_DEFAULT = (
    "https://mcc.admissions.nic.in/Counseling/Report/ViewInstituteProfileDynamic.aspx"
    "?agencyid=140&boardid=140032521"
)

# Two MCC counselling boards — both use the same handler URLs, only the
# boardid query parameter differs. Course label is used for Drive folder names.
LISTINGS: list[tuple[str, str]] = [
    ("PG", "https://mcc.admissions.nic.in/Counseling/Report/"
           "ViewInstituteProfileDynamic.aspx?agencyid=140&boardid=140032521"),
    ("UG", "https://mcc.admissions.nic.in/Counseling/Report/"
           "ViewInstituteProfileDynamic.aspx?agencyid=140&boardid=140012521"),
]

# Direct PDF handler URLs (discovered via diagnostic logging).
# Both are simple GETs parameterized by boardId + InstituteId — no postback needed.
PROFILE_URL_TMPL = (
    "https://mcc.admissions.nic.in/Counseling/Handler/"
    "ViewInstituteProfileDetailsDynamic.ashx?boardId={board}&InstituteId={code}"
)
BOND_URL_TMPL = (
    "https://mcc.admissions.nic.in/Counseling/Handler/"
    "ViewInstituteProfileDynamic.ashx?boardId={board}&InstituteId={code}&Type=BM"
)

# Override via env var (used on Linux/CI). Default is the Windows path you've been using.
STAGING_DIR = Path(
    os.environ.get("MCC_ADMISSIONS_STAGING")
    or "C:/MCC-Admissions-Staging"
)


def _board_id_from_url(listing_url: str) -> str:
    from urllib.parse import urlparse, parse_qs
    qs = parse_qs(urlparse(listing_url).query)
    # Query keys are case-insensitive in practice — try both spellings
    for key in ("boardid", "boardId", "BoardId", "BOARDID"):
        if key in qs and qs[key]:
            return qs[key][0]
    return ""


@dataclass
class InstituteRow:
    code: str
    name: str
    profile_target: str = ""   # __doPostBack target for profile link
    bond_target: str = ""      # __doPostBack target for bond link
    profile_drive_id: str = ""
    bond_drive_id: str = ""
    profile_drive_link: str = ""
    bond_drive_link: str = ""


# --------------------------------------------------------------------------- #
# Postback-target extraction from the initial page HTML
# --------------------------------------------------------------------------- #

POSTBACK_RE = re.compile(r"__doPostBack\('([^']+)','")


def _postback_target(a_tag) -> str:
    if not a_tag:
        return ""
    href = a_tag.get("href", "") or ""
    onclick = a_tag.get("onclick", "") or ""
    for src in (href, onclick):
        m = POSTBACK_RE.search(src)
        if m:
            return m.group(1)
    return ""


def extract_rows_from_html(html: str, log: logging.Logger) -> list[InstituteRow]:
    """Parse the rendered listing page, yield one InstituteRow per table row.
    Walks every <table> in the document and keeps rows whose 2nd or 3rd cell
    looks like an MCC institute code + name. This is selector-agnostic so
    DataTables re-wrapping won't break it.
    """
    soup = BeautifulSoup(html, "html.parser")
    rows: list[InstituteRow] = []
    tables = soup.find_all("table")
    log.info(f"page contains {len(tables)} <table> element(s)")
    for table in tables:
        for tr in table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue
            # Find the first cell that's a 6-digit institute code (like 200101).
            code_idx = -1
            for idx, td in enumerate(tds[:4]):
                txt = td.get_text(" ", strip=True)
                if re.fullmatch(r"\d{5,7}", txt):
                    code_idx = idx
                    break
            if code_idx == -1:
                continue
            code = tds[code_idx].get_text(" ", strip=True)
            # Name is usually the cell right after the code
            name_idx = code_idx + 1
            if name_idx >= len(tds):
                continue
            name = tds[name_idx].get_text(" ", strip=True)
            # Profile and bond link cells come after the name — find them by
            # looking for <a> tags whose postback target contains lnkviewprofile / lnkviewBond
            profile_target = bond_target = ""
            for td in tds[name_idx + 1:]:
                for a in td.find_all("a"):
                    pb = _postback_target(a)
                    if not pb:
                        continue
                    if "lnkviewprofile" in pb.lower():
                        profile_target = pb
                    elif "lnkviewbond" in pb.lower() or "lnkbond" in pb.lower():
                        bond_target = pb
            if not profile_target and not bond_target:
                continue  # this row has no useful links
            rows.append(InstituteRow(
                code=code, name=name,
                profile_target=profile_target,
                bond_target=bond_target,
            ))
    return rows


# --------------------------------------------------------------------------- #
# Robust PDF-capture around a __doPostBack click
# --------------------------------------------------------------------------- #


PDF_CONTENT_TYPES = (
    "application/pdf",
    "application/octet-stream",
    "application/x-download",
    "application/force-download",
    "binary/octet-stream",
)


def _looks_like_pdf_response(resp) -> bool:
    try:
        ct = (resp.headers.get("content-type") or "").lower()
        if any(t in ct for t in PDF_CONTENT_TYPES):
            return True
        url = (resp.url or "").lower()
        if ".pdf" in url:
            return True
        cd = (resp.headers.get("content-disposition") or "").lower()
        if "attachment" in cd or ".pdf" in cd:
            return True
    except Exception:
        pass
    return False


def trigger_postback_and_capture(
    page: Page, target: str, dest: Path, log: logging.Logger, label: str
) -> bool:
    """
    Fire `__doPostBack('<target>', '')` via JavaScript and capture the resulting PDF.

    Server may answer in many shapes, all handled here:
        (a) real browser download event  -> page.on('download')
        (b) inline PDF response bytes     -> page.on('response') with PDF-ish content-type
        (c) navigation to a PDF URL       -> framenavigated to URL containing .pdf
        (d) new tab/popup with PDF        -> context.on('page')
    Returns True on success.
    """
    captured = {"bytes": None, "dl": None, "popup": None, "nav_url": None}
    response_log: list[tuple] = []  # (status, ct, url, len)

    def on_response(resp):
        try:
            ct = (resp.headers.get("content-type") or "").lower()
            response_log.append((resp.status, ct[:60], resp.url[:160], 0))
            if _looks_like_pdf_response(resp) and resp.status < 400 and not captured["bytes"]:
                body = resp.body()
                if body[:5] == b"%PDF-" or len(body) > 1000:
                    captured["bytes"] = body
        except Exception:
            pass

    def on_download(d):
        captured["dl"] = d

    def on_popup(p):
        captured["popup"] = p

    def on_framenav(frame):
        try:
            if frame == page.main_frame:
                u = (frame.url or "").lower()
                if ".pdf" in u or "viewpdf" in u or "showpdf" in u or "downloadpdf" in u:
                    captured["nav_url"] = frame.url
        except Exception:
            pass

    page.on("response", on_response)
    page.on("download", on_download)
    page.context.on("page", on_popup)
    page.on("framenavigated", on_framenav)

    try:
        page.evaluate(f"__doPostBack({target!r}, '')")
    except Exception as e:
        log.warning(f"[{label}] __doPostBack evaluate error: {e}")

    # Wait up to 25s for any capture channel
    deadline = time.time() + 25
    while time.time() < deadline:
        if captured["dl"] or captured["bytes"] or captured["popup"] or captured["nav_url"]:
            try:
                page.wait_for_timeout(800)
            except Exception:
                pass
            break
        try:
            page.wait_for_timeout(250)
        except Exception:
            time.sleep(0.25)

    try:
        page.remove_listener("response", on_response)
        page.remove_listener("download", on_download)
        page.context.remove_listener("page", on_popup)
        page.remove_listener("framenavigated", on_framenav)
    except Exception:
        pass

    # Try save preference: native download > inline bytes > popup PDF > navigation URL
    try:
        if captured["dl"]:
            captured["dl"].save_as(str(dest))
            if dest.exists() and dest.stat().st_size > 0 and _is_valid_pdf(dest):
                return True

        if captured["bytes"]:
            dest.write_bytes(captured["bytes"])
            if _is_valid_pdf(dest):
                return True

        if captured["popup"]:
            popup = captured["popup"]
            try:
                popup.wait_for_load_state("domcontentloaded", timeout=10000)
            except Exception:
                pass
            popup_url = popup.url
            log.info(f"[{label}] popup opened: {popup_url}")
            # If popup has a PDF URL, fetch via the context's request session (cookies preserved)
            if popup_url and popup_url != "about:blank":
                try:
                    resp = page.context.request.get(popup_url)
                    body = resp.body()
                    if body[:5] == b"%PDF-":
                        dest.write_bytes(body)
                        return True
                except Exception as e:
                    log.warning(f"[{label}] popup fetch failed: {e}")
            try:
                popup.close()
            except Exception:
                pass

        if captured["nav_url"]:
            log.info(f"[{label}] navigated to: {captured['nav_url']}")
            try:
                resp = page.context.request.get(captured["nav_url"])
                body = resp.body()
                if body[:5] == b"%PDF-":
                    dest.write_bytes(body)
                    return True
            except Exception as e:
                log.warning(f"[{label}] nav fetch failed: {e}")
    except Exception as e:
        log.warning(f"[{label}] save failed: {e}")

    # Diagnostic dump
    log.warning(f"[{label}] capture failed. Last responses (most recent first):")
    for status, ct, url, _ in response_log[-12:][::-1]:
        log.warning(f"    {status} {ct:<60} {url}")
    return False


def _is_valid_pdf(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False


def fetch_pdf_via_session(page: Page, url: str, dest: Path,
                          log: logging.Logger, label: str) -> bool:
    """Fetch a PDF directly via the page context's request session
    (cookies + ASP.NET session preserved). Far simpler than the postback
    capture dance — used after we discovered the .ashx handler URLs."""
    try:
        resp = page.context.request.get(url, timeout=60000)
    except Exception as e:
        log.warning(f"[{label}] request failed: {e}")
        return False
    if not resp.ok:
        log.warning(f"[{label}] HTTP {resp.status} — skipping")
        return False
    try:
        body = resp.body()
    except Exception as e:
        log.warning(f"[{label}] body read failed: {e}")
        return False
    if not body or body[:5] != b"%PDF-":
        snippet = body[:80] if body else b""
        log.warning(f"[{label}] response is not a PDF (first bytes: {snippet!r})")
        return False
    try:
        dest.write_bytes(body)
    except Exception as e:
        log.warning(f"[{label}] disk write failed: {e}")
        return False
    return dest.stat().st_size > 0


# --------------------------------------------------------------------------- #
# Filter cascade for the listing page (Type -> Institutes -> Program -> Submit)
# --------------------------------------------------------------------------- #


def _wait_for_select_option(page: Page, select_id: str, value: str,
                            log: logging.Logger, timeout_ms: int = 25000) -> bool:
    """Poll the DOM until <select id=...> has an <option value=...> available.
    Necessary because ASP.NET partial postbacks (UpdatePanel) sometimes complete
    networkidle before the DOM mutation lands."""
    try:
        page.wait_for_function(
            """(args) => {
                const sel = document.getElementById(args.id);
                if (!sel || sel.options.length === 0) return false;
                return Array.from(sel.options).some(o => o.value === args.value);
            }""",
            arg={"id": select_id, "value": value},
            timeout=timeout_ms,
        )
        return True
    except PWTimeout:
        log.warning(f"    timed out waiting for #{select_id} to have option {value!r}")
        return False


def _set_native_select(page: Page, select_id: str, value: str, log: logging.Logger,
                       wait: bool = True) -> bool:
    """Set a hidden native <select>'s value via JS (Chosen.js hides the element).
    If `wait` is True, polls until the option is present in the DOM before setting.
    Returns True if the value was actually selectable."""
    if wait and not _wait_for_select_option(page, select_id, value, log):
        # Try once anyway in case the option just appeared
        pass
    result = page.evaluate(
        """(args) => {
            const sel = document.getElementById(args.id);
            if (!sel) return { ok: false, reason: 'missing' };
            const opt = Array.from(sel.options).find(o => o.value === args.value);
            if (!opt) {
                return {
                    ok: false, reason: 'option_missing',
                    n_options: sel.options.length,
                    sample: Array.from(sel.options).slice(0, 4).map(o => ({v: o.value, t: o.text}))
                };
            }
            sel.value = args.value;
            // Notify Chosen.js / any listeners
            sel.dispatchEvent(new Event('change', {bubbles: true}));
            return { ok: true };
        }""",
        {"id": select_id, "value": value},
    )
    if result.get("ok"):
        log.info(f"    filter set: #{select_id} = {value!r}")
        return True
    log.warning(f"    filter FAILED: #{select_id} = {value!r} -> {result}")
    return False


def _trigger_postback(page: Page, target: str, log: logging.Logger):
    """Fire __doPostBack(target, '') and wait for the partial postback to fully complete.
    Listens for the actual response to the __EVENTTARGET POST so we don't return early
    before the DOM mutation lands."""
    try:
        with page.expect_response(
            lambda r: "ViewInstituteProfileDynamic.aspx" in r.url
                      and r.request.method == "POST",
            timeout=30000,
        ):
            page.evaluate(f"__doPostBack({target!r}, '')")
    except PWTimeout:
        log.warning(f"    no POST response for postback {target!r} within 30s")
    except Exception as e:
        log.warning(f"    postback evaluate error: {e}")
    # Settle: networkidle gives any follow-up requests a chance, then small grace
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PWTimeout:
        pass
    page.wait_for_timeout(400)


def expand_datatables_to_all(page: Page, log: logging.Logger):
    """Force every DataTables grid on the page to render ALL rows in the DOM.
    DataTables paginates client-side (10 rows per page by default). To extract
    every postback target we need every row materialised at once."""
    log.info("Expanding DataTables to render all rows ...")
    try:
        result = page.evaluate(
            """
            (() => {
                const out = {via_dt_api: 0, via_length_select: 0, errors: []};
                // Path 1: DataTables jQuery API
                try {
                    if (typeof jQuery !== 'undefined' && jQuery.fn.dataTable) {
                        jQuery('table').each(function() {
                            try {
                                if (jQuery.fn.dataTable.isDataTable(this)) {
                                    jQuery(this).DataTable().page.len(-1).draw();
                                    out.via_dt_api++;
                                }
                            } catch (e) { out.errors.push(String(e)); }
                        });
                    }
                } catch (e) { out.errors.push('api:'+e); }
                // Path 2: change the visible page-length <select>
                try {
                    document.querySelectorAll('select[name$="_length"]').forEach(s => {
                        const opt = Array.from(s.options).find(o => o.value === '-1')
                                 || Array.from(s.options).find(o => /all/i.test(o.text))
                                 || s.options[s.options.length - 1];
                        if (opt) {
                            s.value = opt.value;
                            s.dispatchEvent(new Event('change', {bubbles: true}));
                            out.via_length_select++;
                        }
                    });
                } catch (e) { out.errors.push('len_select:'+e); }
                return out;
            })()
            """
        )
        log.info(f"    expansion result: {result}")
    except Exception as e:
        log.warning(f"    DataTables expansion threw: {e}")
    page.wait_for_timeout(2500)


def populate_filters_and_submit(page: Page, log: logging.Logger):
    """Walk the cascaded filters (Type -> Institutes -> Program) and click Submit."""
    log.info("Populating filter cascade (Type -> Institutes -> Program -> Submit) ...")

    # 1. Type of Institute — 'All' is a fixed option in the initial HTML
    if _set_native_select(page, "ctl00_ContentPlaceHolder1_ddlInstType", "All", log):
        _trigger_postback(page, "ctl00$ContentPlaceHolder1$ddlInstType", log)

    # 2. Institutes — 'All' appears only after Type postback populates this dropdown
    if _set_native_select(page, "ctl00_ContentPlaceHolder1_ddlInstitutes", "All", log):
        _trigger_postback(page, "ctl00$ContentPlaceHolder1$ddlInstitutes", log)
    else:
        log.warning("Institutes 'All' option not present; continuing")

    # 3. Program — no onchange on this one; just set and leave
    _set_native_select(page, "ctl00_ContentPlaceHolder1_ddlprogram", "All", log)

    # 4. Click Submit
    log.info("    clicking Submit ...")
    try:
        page.click("#ctl00_ContentPlaceHolder1_btnsubmit", timeout=10000)
    except Exception as e:
        log.warning(f"    native click failed ({e}); using JS fallback")
        page.evaluate(
            "document.getElementById('ctl00_ContentPlaceHolder1_btnsubmit').click()"
        )
    for state in ("domcontentloaded", "networkidle"):
        try:
            page.wait_for_load_state(state, timeout=30000)
        except PWTimeout:
            pass
    page.wait_for_timeout(1500)


# --------------------------------------------------------------------------- #
# Excel index writer
# --------------------------------------------------------------------------- #


def write_and_upload_excel(rows: list[InstituteRow], drive: DriveUploader,
                           log: logging.Logger, course: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from googleapiclient.http import MediaFileUpload

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course} Institute Profiles"
    headers = ["Institute Code", "Institute Name", "Profile PDF", "Bond PDF"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([r.code, r.name, "", ""])
        row_n = ws.max_row
        if r.profile_drive_link:
            cell = ws.cell(row=row_n, column=3, value="View profile")
            cell.hyperlink = r.profile_drive_link
            cell.font = Font(color="0000EE", underline="single")
        if r.bond_drive_link:
            cell = ws.cell(row=row_n, column=4, value="View bond")
            cell.hyperlink = r.bond_drive_link
            cell.font = Font(color="0000EE", underline="single")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"

    course_staging = STAGING_DIR / course
    course_staging.mkdir(parents=True, exist_ok=True)
    file_name = f"Institute_Profile_Index_{course}.xlsx"
    out_path = course_staging / file_name
    wb.save(out_path)
    log.info(f"[{course}] Excel saved locally: {out_path}")

    # Upload to Drive under Institute Profiles/<course>/
    folder_id = drive._ensure_folder(("Institute Profiles", course))
    media = MediaFileUpload(
        str(out_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    existing = drive.svc.files().list(
        q=f"'{folder_id}' in parents and name='{file_name}' and trashed=false",
        fields="files(id)", pageSize=1,
    ).execute().get("files", [])
    if existing:
        updated = drive.svc.files().update(
            fileId=existing[0]["id"], media_body=media,
            body={"name": file_name}, fields="id",
        ).execute()
        log.info(f"[{course}] Excel updated on Drive: {updated['id']}")
        return updated["id"]
    created = drive.svc.files().create(
        media_body=media,
        body={"name": file_name, "parents": [folder_id]},
        fields="id",
    ).execute()
    log.info(f"[{course}] Excel uploaded to Drive: {created['id']}")
    return created["id"]


# --------------------------------------------------------------------------- #
# Orchestrator
# --------------------------------------------------------------------------- #


def run(listing_url: str, limit: int | None, headless: bool,
        course: str = "PG", drive: DriveUploader | None = None,
        log: logging.Logger | None = None):
    cfg = load_config()
    if log is None:
        logs_dir = Path(cfg["paths"]["logs"]).resolve()
        log = setup_logging(logs_dir)
    log.info(f"=== MCC admissions pipeline [{course}] ===")
    log.info(f"URL: {listing_url}")
    log.info(f"limit={limit} headless={headless}")

    course_staging = STAGING_DIR / course
    course_staging.mkdir(parents=True, exist_ok=True)
    if drive is None:
        drive = DriveUploader(cfg, log)
    profile_folder = ["Institute Profiles", course, "Profile"]
    bond_folder = ["Institute Profiles", course, "Bond"]

    with sync_playwright() as p:
        # Disable the "HeadlessChrome" fingerprint and other automation flags —
        # many gov.nic.in sites reject headless Chromium outright.
        browser = p.chromium.launch(
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
        ctx = browser.new_context(
            accept_downloads=True,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
            locale="en-IN",
        )
        # Remove the navigator.webdriver flag that headless Chromium sets
        ctx.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = ctx.new_page()

        log.info("Loading listing page ...")
        page.goto(listing_url, wait_until="domcontentloaded", timeout=90000)
        # Wait for the Chosen.js scripts and ASP.NET form scripts to register.
        try:
            page.wait_for_function(
                "typeof __doPostBack === 'function'", timeout=15000
            )
        except PWTimeout:
            log.warning("__doPostBack never registered; continuing anyway")
        page.wait_for_timeout(1000)

        # Drive the dependent dropdowns (Type -> Institutes -> Program) and Submit.
        try:
            populate_filters_and_submit(page, log)
        except Exception as e:
            log.exception(f"filter cascade failed: {e}")

        # Expand DataTables so every row materialises in the DOM (default is 10).
        try:
            expand_datatables_to_all(page, log)
        except Exception as e:
            log.exception(f"DataTables expansion failed: {e}")

        def _try_extract() -> list[InstituteRow]:
            try:
                return extract_rows_from_html(page.content(), log)
            except Exception as e:
                log.warning(f"extract failed (will retry): {e}")
                return []

        # After Submit, the grid renders client-side via DataTables — give it
        # a few seconds and retry extraction.
        rows: list[InstituteRow] = []
        deadline = time.time() + 30
        while time.time() < deadline:
            rows = _try_extract()
            if rows:
                break
            page.wait_for_timeout(1500)

        log.info(f"Discovered {len(rows)} institute rows")
        if not rows:
            # Dump evidence for debugging
            dump_html = STAGING_DIR / "debug_page.html"
            dump_png = STAGING_DIR / "debug_page.png"
            try:
                dump_html.write_text(page.content(), encoding="utf-8")
                page.screenshot(path=str(dump_png), full_page=True)
                log.error(f"No rows discovered. Dumped HTML -> {dump_html}")
                log.error(f"Screenshot -> {dump_png}")
            except Exception as e:
                log.error(f"No rows discovered and dump failed: {e}")
            browser.close()
            return

        if limit:
            rows = rows[:limit]
            log.info(f"Limiting to first {len(rows)} rows (smoke-test)")

        # Resolve boardId once — it parameterizes every PDF URL
        board_id = _board_id_from_url(listing_url)
        if not board_id:
            log.error(f"Could not extract boardid from listing URL: {listing_url}")
            browser.close()
            return
        log.info(f"Using boardId={board_id} for direct PDF handler URLs")

        for i, rec in enumerate(rows, start=1):
            log.info(f"[{i}/{len(rows)}] {rec.code} — {rec.name[:70]}")

            safe_name = sanitize_name(f"{rec.code} {rec.name}", 110) + ".pdf"

            # PROFILE — direct .ashx handler fetch
            profile_url = PROFILE_URL_TMPL.format(board=board_id, code=rec.code)
            dest = course_staging / f"profile_{rec.code}.pdf"
            if fetch_pdf_via_session(page, profile_url, dest, log, f"profile/{rec.code}"):
                try:
                    ur = PdfRecord(
                        url=profile_url,
                        breadcrumb=profile_folder,
                        heading=rec.name,
                        link_text=f"{rec.code} {rec.name}",
                        local_path=str(dest),
                    )
                    rec.profile_drive_id = drive.upload(ur, profile_folder, safe_name)
                    rec.profile_drive_link = (
                        f"https://drive.google.com/file/d/{rec.profile_drive_id}/view"
                    )
                    dest.unlink(missing_ok=True)
                except Exception as e:
                    log.exception(f"profile upload failed for {rec.code}: {e}")
            else:
                log.warning(f"[profile/{rec.code}] fetch failed")

            # BOND — direct .ashx handler fetch
            bond_url = BOND_URL_TMPL.format(board=board_id, code=rec.code)
            dest = course_staging / f"bond_{rec.code}.pdf"
            if fetch_pdf_via_session(page, bond_url, dest, log, f"bond/{rec.code}"):
                try:
                    ur = PdfRecord(
                        url=bond_url,
                        breadcrumb=bond_folder,
                        heading=rec.name,
                        link_text=f"{rec.code} {rec.name}",
                        local_path=str(dest),
                    )
                    rec.bond_drive_id = drive.upload(ur, bond_folder, safe_name)
                    rec.bond_drive_link = (
                        f"https://drive.google.com/file/d/{rec.bond_drive_id}/view"
                    )
                    dest.unlink(missing_ok=True)
                except Exception as e:
                    log.exception(f"bond upload failed for {rec.code}: {e}")
            else:
                log.info(f"[bond/{rec.code}] no bond available (or fetch failed)")

        browser.close()

    # Build Excel index and upload
    log.info(f"[{course}] Building Excel index ...")
    write_and_upload_excel(rows, drive, log, course)

    uploaded_p = sum(1 for r in rows if r.profile_drive_id)
    uploaded_b = sum(1 for r in rows if r.bond_drive_id)
    log.info(f"[{course}] Done. institutes={len(rows)} "
             f"profile_uploaded={uploaded_p} bond_uploaded={uploaded_b}")

    # Diff against last-run snapshot to identify NEW institutes (codes not seen
    # before). Snapshot lives at C:/MCC-Admissions-Staging/<course>/last_state.json.
    state_path = course_staging / "last_state.json"
    prev_codes: set[str] = set()
    if state_path.exists():
        try:
            import json
            prev_codes = set(json.loads(state_path.read_text("utf-8")).get("codes", []))
        except Exception as e:
            log.warning(f"[{course}] could not read previous state: {e}")
    current_codes = {r.code for r in rows}
    new_codes = current_codes - prev_codes
    new_institutes = [
        {
            "code": r.code,
            "name": r.name,
            "profile_link": r.profile_drive_link,
            "bond_link": r.bond_drive_link,
        }
        for r in rows if r.code in new_codes
    ]
    # Save new state (unless this was a partial smoke test — heuristic: full row count)
    if not limit:
        try:
            import json
            state_path.write_text(
                json.dumps({"codes": sorted(current_codes),
                            "saved_at": time.strftime("%Y-%m-%dT%H:%M:%S")},
                           ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as e:
            log.warning(f"[{course}] could not save state: {e}")
    log.info(f"[{course}] new institutes since last run: {len(new_institutes)}")

    return {
        "pipeline": f"admissions_{course.lower()}",
        "course": course,
        "institutes": len(rows),
        "profile_uploaded": uploaded_p,
        "bond_uploaded": uploaded_b,
        "new_institutes": new_institutes,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", default=None,
                    help="Override listing URL (otherwise uses LISTINGS for the chosen course)")
    ap.add_argument("--course", choices=["PG", "UG", "BOTH"], default="BOTH",
                    help="Which counselling board to crawl (default: both)")
    ap.add_argument("--limit", type=int, default=None,
                    help="Stop after N institutes per course (smoke test)")
    ap.add_argument("--headed", action="store_true",
                    help="Run Chromium with a visible window (for debugging)")
    args = ap.parse_args()

    cfg = load_config()
    logs_dir = Path(cfg["paths"]["logs"]).resolve()
    log = setup_logging(logs_dir)
    drive = DriveUploader(cfg, log)

    if args.url:
        # Manual override — single run
        course = args.course if args.course != "BOTH" else "PG"
        run(args.url, args.limit, headless=not args.headed,
            course=course, drive=drive, log=log)
        return

    # Loop over LISTINGS, respecting --course filter
    targets = [(c, u) for (c, u) in LISTINGS
               if args.course == "BOTH" or c == args.course]
    if not targets:
        log.error(f"No listings match --course={args.course}")
        return

    summaries = []
    for course, url in targets:
        log.info(f"\n{'='*60}\n  Starting course: {course}\n{'='*60}")
        try:
            summary = run(url, args.limit, headless=not args.headed,
                          course=course, drive=drive, log=log)
            if summary:
                summaries.append(summary)
        except Exception as e:
            log.exception(f"[{course}] pipeline crashed: {e}")

    log.info("\n" + "="*60)
    log.info("  ALL COURSES COMPLETE")
    log.info("="*60)
    for s in summaries:
        log.info(f"  [{s['course']}] institutes={s['institutes']} "
                 f"profile_uploaded={s['profile_uploaded']} "
                 f"bond_uploaded={s['bond_uploaded']}")


if __name__ == "__main__":
    main()
